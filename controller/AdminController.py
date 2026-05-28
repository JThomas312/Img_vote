#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep  9 19:33:36 2025

@author: jacques
"""

#general imports
from bcrypt import gensalt, hashpw

from openpyxl import Workbook
from os import getcwd
import os.path
from os import remove
from math import ceil
from zipfile import ZipFile
from shutil import rmtree
from shutil import copyfile
from re import match

#enable imports from local modules
from pathlib import Path
import sys
path_root = Path(__file__).parents[2]
sys.path.append(str(path_root))

#local modules
from img_vote.utilities.useful import generate_password
from img_vote.utilities.useful import format_r_friendly
from img_vote.utilities.useful import sanitize

from img_vote.Models.ViewModels import CriterionEditingViewModel, CategoryConfigurationViewModel
from img_vote.Models.ViewModels import CategoryEditingViewModel, PrerequisiteEditingViewModel, UploadStatusViewModel, ReviewerDistributionViewmodel

#user related
from img_vote.dal.MasterDal import get_reviewer_by_id, get_reviewer_by_login, count_all_reviewers, create_reviewer 
from img_vote.dal.MasterDal import delete_reviewer_by_id, update_password, clear_non_admin_users

#case related
from img_vote.dal.MasterDal import count_all_cases, extract_all_data, create_all_cases, clear_all_cases

#answer related
from img_vote.dal.MasterDal import create_all_answers, create_user_answers

#category related
from img_vote.dal.MasterDal import get_category_by_id, categories_with_criteria, category_with_criteria_and_prerequisites
from img_vote.dal.MasterDal import at_least_one_other_mandatory_category, at_least_one_mandatory_category, tutorial_category_exists
from img_vote.dal.MasterDal import categories_without_name, mandatory_categories_with_prerequisites, optional_categories_without_prerequisites
from img_vote.dal.MasterDal import categories_without_criteria, malignant_categories_without_gold_standard, other_gold_standard_exists
from img_vote.dal.MasterDal import gold_standard_exists, get_gold_standards, gold_standard_in_wrong_category, new_empty_category
from img_vote.dal.MasterDal import update_category_value, erase_category, clear_all_categories, get_na_tutorial_categories

#prerequisite related
from img_vote.dal.MasterDal import new_prerequisite, delete_prerequisite, delete_category_prerequisite
from img_vote.dal.MasterDal import delete_prerequisite_from_category_criteria, delete_prerequisite_from_criterion, clear_all_prerequisites

#criterion related
from img_vote.dal.MasterDal import create_criterion, update_criterion, update_criterion_malignancy, update_criteria_path
from img_vote.dal.MasterDal import clear_malignant_criteria_in_non_malignant_category, erase_criterion
from img_vote.dal.MasterDal import erase_category_criteria, create_na_criteria, create_trust_criteria, remove_na_criteria
from img_vote.dal.MasterDal import remove_trust_criteria, create_all_answer_to_criterion, clear_all_answers
from img_vote.dal.MasterDal import create_user_answer_to_criterion, clear_all_criteria


def find_name_and_login(userId):
    usr = get_reviewer_by_id(userId)
    usrName = usr.name
    usrLogin = usr.login
    return (usrName, usrLogin)


def create_user(login, name, admin, status, full_review):
    existing = get_reviewer_by_login(login)
    full = full_review and not admin
    with open(os.path.join(getcwd(), 'persistence', 'distribution.txt'), 'r', encoding="utf-8") as fr:
        distribution = fr.readline().removesuffix('\n')
        case_per_r = fr.readline().removesuffix('\n')
        percentage = fr.readline().removesuffix('\n')
    if ((not admin) and (status == 'ended')):
        raise Exception('While the study is ended only administrator users can be created')
    elif not full and distribution == 'n per case' and status in ['ready', 'ongoing', 'paused']:
        raise Exception('You can now only create full reviewers with your chosen distribution')
    elif existing != None:
        raise Exception('User already exists with login ' + login)
    else:
        password = generate_password()
        s = gensalt()
        hashPass = hashpw(password.encode('utf-8'), s).decode('utf-8')
        revId = create_reviewer(name, login, hashPass, admin, full)
        if not admin and status in ['ready', 'ongoing', 'paused']:
            case_per_rev = compute_case_per_rev(full, distribution, case_per_r, percentage)
            create_user_answers(revId, case_per_rev)
            create_user_answer_to_criterion(revId)
    return password

def categories_for_editing():
    categoriesDMs = categories_with_criteria()
    categoriesVMs = []
    for categoryDM in categoriesDMs:
        currentCategoryVM = CategoryConfigurationViewModel(categoryDM.catId, categoryDM.name, categoryDM.catType, categoryDM.hasTrust, categoryDM.hasTutorial, categoryDM.hasNA, categoryDM.optional, categoryDM.hasGoldStandard, categoryDM.hasMalignancy)
        for crit in categoryDM.criteria:
            currentCategoryVM.criteria.append(CriterionEditingViewModel(crit[0], crit[1]))
        categoriesVMs.append(currentCategoryVM)
    return categoriesVMs   

def category_for_editing(catId):
    categoryDM = category_with_criteria_and_prerequisites(catId)
    categoryVM = CategoryEditingViewModel(categoryDM.catId, categoryDM.name, categoryDM.catType, categoryDM.hasTrust, categoryDM.hasTutorial, categoryDM.hasNA, categoryDM.optional, categoryDM.hasGoldStandard, categoryDM.hasMalignancy)
    
    for criterion in categoryDM.criteria:
        categoryVM.criteria.append(CriterionEditingViewModel(criterion[0], criterion[1], criterion[2]))
    
    for prerequisite in categoryDM.prerequisites:
        categoryVM.prerequisites.append(PrerequisiteEditingViewModel(prerequisite[0], prerequisite[1]))

    return categoryVM 

def create_empty_category():
    newId = new_empty_category()
    return newId

def update_category(cat_id, value, parameter):
    val = value
    param = parameter
    #this way if either the DAL of front changes parameter, only this needs to be updated --> no dependency between front and DAL
    if parameter == 'name':
        if value == '' or not sanitize(value):
            return
        param = 'name'
    if parameter == 'type':
        val = int(value)
        param = 'type'
    if parameter == 'tutorial':
        param = 'tutorial'
    if parameter == 'trust':
        param = 'trust'
    if parameter == 'na':
        param = 'na'
    if parameter == 'optional':
        param = 'optional'
    if parameter == 'gold_standard':
        param = 'gold_standard'
    if parameter == 'malignancy':
        param ='malignancy'
    if (value == 'true'):
        val = True
    if (value == 'false'):
        val = False
        
    update_category_value(cat_id, val, param)
    return

def save_criterion(cat_id, name, malignancy):
    
    if not sanitize(name) or name == '':
        return
    
    category_name = get_category_by_id(cat_id).name
    tutorial_slide_path = os.path.join(getcwd(), 'data', 'tutorial_data', category_name, name)
    mal = malignancy == 'true'
    crit_id = create_criterion(name, tutorial_slide_path, cat_id, False, mal)
    
    return crit_id

def change_criterion(cat_id, crit_id, name, malignancy, action):
    if action == 'remove':
        delete_prerequisite_from_criterion(crit_id)
        erase_criterion(crit_id)    
    if not sanitize(name) or name == '':
        return
    if action == 'edit':
        mal = malignancy == 'true'
        update_criterion(crit_id, name, mal)

def save_criterion_malignancy(crit_id, malignancy):
    mal = malignancy == 'true'
    update_criterion_malignancy(crit_id, mal)

def save_prerequisite(cat_id, name):
    if not sanitize(name) or name == '':
        return
    crit_id = new_prerequisite(cat_id, name)
    
    return crit_id

def change_prerequisite(cat_id, crit_id, name, action):
    if not sanitize(name) or name == '':
        return
    if action == 'remove':
        delete_prerequisite(cat_id, crit_id)    
    if action == 'edit':
        delete_prerequisite(cat_id, crit_id) 
        return new_prerequisite(cat_id, name)
        

def delete_category(cat_id):
    delete_category_prerequisite(cat_id)
    delete_prerequisite_from_category_criteria(cat_id)
    erase_category_criteria(cat_id)
    erase_category(cat_id)

def delete_criterion(crit_id):
    delete_prerequisite_from_criterion(crit_id)
    erase_criterion(crit_id)

def check_category(cat_id, form_answers):

    if (not 'name' in form_answers) or (form_answers['name'] == '') or (not sanitize(form_answers['name'])):
        return 'Invalid category : your category must have a valid name'

    update_category(cat_id, form_answers['name'], 'name')

    if not 'type' in form_answers:
        return 'Invalid category : choosing type is mandatory'
    
    update_category(cat_id, form_answers['type'], 'type')
    
    prerequisites_ok = True
    if 'optional' in form_answers and form_answers['optional'] == '1':
        prerequisites_ok = False
    
    nb_answers = 0
    
    for ans in form_answers:
        if ans.find('prerequisiteField') != -1:
            val = form_answers[ans]
            if val == '':
                return 'Invalid category : one of your answers is unnamed'
            if not sanitize(val):
                return 'invalid prerequisite : ' + val + ' is not a valid name'
            prerequisites_ok = True
        if ans.find('criterionField') != -1:
            val = form_answers[ans]
            if val == '':
                return 'Invalid category : one of your answers is unnamed'
            if not sanitize(val):
                return 'invalid answer : ' + val + ' is not a valid name'
            nb_answers += 1
    
    if not prerequisites_ok:
        return 'Invalid category : an optional category needs at least one prerequisite'
    
    if form_answers['type'] == '2' and nb_answers < 2:
        return 'Invalid category : this type of category needs at least two answers'
    
    if nb_answers == 0:
        return 'Invalid category : every category needs at least one answer'
    
    return None

def check_categories():
    
    errors = []
    
    incorrect_categories = []
    unnamed_categories = categories_without_name()
    if len(unnamed_categories) > 0:
        errors.append('Some of your categories are still unnamed')
        incorrect_categories.extend(unnamed_categories)
    if not at_least_one_mandatory_category():
        errors.append('All of your categories are optional, you need at least one mandatory category')
    
    gold_standards = get_gold_standards()
    if len(gold_standards) > 1:
        errors.append('Several categories are marked as having a gold standard but only one is allowed per study')
        incorrect_categories.extend(gold_standards)
    
    wrong_gold_standards = gold_standard_in_wrong_category()
    if len(wrong_gold_standards) > 0:
        errors.append('''Your gold standard category is not of one-of type, please change it's type or remove the gold standard''')
        incorrect_categories.extend(wrong_gold_standards)
    
    mandatory_categories = mandatory_categories_with_prerequisites()
    if len(mandatory_categories) > 0:
        errors.append('Some mandatory categories have prerequisites, please remove them or make them optional')
        incorrect_categories.extend(mandatory_categories)
    
    optional_categories = optional_categories_without_prerequisites()
    if len(optional_categories) > 0:
        errors.append('Some optional categories have no prerequisites, please add at least one or make them mandatory')
        incorrect_categories.extend(optional_categories)
    
    no_criteria = categories_without_criteria()
    if len(no_criteria) > 0:
        errors.append('Some categories have no possible answer, please add at least one or delete the category')
        incorrect_categories.extend(no_criteria)
    
    malignant_categories = malignant_categories_without_gold_standard()        
    if len(malignant_categories) > 0:
        errors.append('Some categories are marked as having malignancy but not gold standard, please check them')
        incorrect_categories.extend(malignant_categories)

    unique_incorrect_categories = list(dict.fromkeys(incorrect_categories))
    
    for i in range(len(unique_incorrect_categories)):
        (identifier, name) = unique_incorrect_categories[i]
        if bool(match('^[\\s]*$', name)):
            unique_incorrect_categories[i] = (identifier, 'unnamed category')
    
    if len(errors) > 0 or len(unique_incorrect_categories) > 0:
        return (errors, unique_incorrect_categories)

    clear_malignant_criteria_in_non_malignant_category()
    
    create_na_criteria()

    update_criteria_path(os.path.join(getcwd(), 'data', 'tutorial_data'))
    
    create_trust_criteria()

    return ([], [])

def categories_rollback():
    remove_na_criteria()
    remove_trust_criteria()
    remove_case_images()
    remove_tutorial_images()
    remove_case_data()

def optional_category_allowed(categoryId):
    allowed = at_least_one_other_mandatory_category(categoryId)
    return allowed

def gold_standard_category_allowed(categoryId):
    allowed = not other_gold_standard_exists(categoryId)
    return allowed

def upload_status():
    VM = UploadStatusViewModel()
    VM.case_images_uploaded = os.path.exists(os.path.join(getcwd(), 'uploads', 'case_images.zip'))
    VM.tutorial_images_needed = tutorial_category_exists()
    VM.tutorial_images_uploaded = os.path.exists(os.path.join(getcwd(), 'uploads', 'tutorial_images.zip'))
    VM.case_data_needed = gold_standard_exists()
    VM.case_data_uploaded = os.path.exists(os.path.join(getcwd(), 'uploads', 'case_data'))
    return VM

def unzip_and_move(path, version):
    
    if version == 'case':
        extract_path = os.path.join(getcwd(), 'data', 'Img_data')
    elif version == 'tutorial':
        extract_path = os.path.join(getcwd(), 'data', 'tutorial_data')
    else:
        return 'something went wrong'
    
    try:
        zippy = ZipFile(path)
        zippy.extractall(extract_path)
    except Exception as e:
        return e
    
    if version == 'tutorial':
        #all one of categories with tutorials and n/a enabled need the template n/a tutorial
        na_categories = get_na_tutorial_categories()
        
        for na_category in na_categories:
            copyfile(os.path.join(getcwd(), 'data', 'na.jpg'), os.path.join(getcwd(), 'data', 'tutorial_data', na_category.name, 'na.jpg'))
    
    return None

def move(ogpath, newpath):
    copyfile(ogpath, newpath)

def remove_case_images():
    upload_path = os.path.join(getcwd(), 'uploads', 'case_images.zip')
    if os.path.exists(upload_path):
        remove(upload_path)
    unziped_path = os.path.join(getcwd(), 'data', 'Img_data')
    if os.path.exists(unziped_path):
        rmtree(unziped_path)

def remove_tutorial_images():
    upload_path = os.path.join(getcwd(), 'uploads', 'tutorial_images.zip')
    if os.path.exists(upload_path):
        remove(upload_path)
    unziped_path = os.path.join(getcwd(), 'data', 'tutorial_data')
    if os.path.exists(unziped_path):
        rmtree(unziped_path)

def remove_case_data():
    upload_path = os.path.join(getcwd(), 'uploads', 'case_data')
    if os.path.exists(upload_path):
        remove(upload_path)
    for filename in os.listdir(os.path.join(getcwd(), 'data')):
        if bool(match('case_data', filename)):
            remove(os.path.join(getcwd(), 'data', filename))

def check_uploads_and_create_cases():
    problems = create_all_cases()
    if problems != None:
        if problems[0] == 'file name discrepancy':
            return 'Error : file ' + str(problems[1])   + ' does not correspond to case ' + problems[2] + ' in data file, please check your data and upload it again'
        if problems[0] == 'gold standard name discrepancy':
            return 'Error : name ' + str(problems[1]) + ' was encountered in your gold standard spreadsheet but is not an answer in your gold standard category'
    return problems

def upload_rollback():
    clear_all_cases()

def data_for_distribution():
    nb_cases = count_all_cases()
    nb_standard_reviewers = count_all_reviewers(False)
    nb_full_reviewers = count_all_reviewers(True)
    
    distributionVM = ReviewerDistributionViewmodel(nb_cases, nb_standard_reviewers + nb_full_reviewers, nb_full_reviewers)
    
    return distributionVM

def handle_distribution(method, r_per_case, case_per_r, percentage):

    rev_per_case= compute_rev_per_case(method, r_per_case, case_per_r, percentage)

    create_all_answers(rev_per_case)    
    create_all_answer_to_criterion()
    
    with open(os.path.join(getcwd(), 'persistence', 'distribution.txt'), 'w', encoding="utf-8") as fw:
        fw.writelines([method, '\n', case_per_r, '\n', percentage])
    
    return None

def distribution_rollback():
    clear_all_answers()

def compute_rev_per_case(method, r_per_case, case_per_r, percentage):
    nb_cases = count_all_cases()
    nb_standard_reviewers = count_all_reviewers(False)
    nb_full_reviewers = count_all_reviewers(True)
    
    if method == 'all for all' or nb_standard_reviewers == 0:
        rev_per_case = nb_standard_reviewers
    
    if method == 'n per case':
        rev_per_case = int(r_per_case) - nb_full_reviewers
        if rev_per_case > nb_standard_reviewers:
            rev_per_case = nb_standard_reviewers
 
    if method == 'n per reviewer':
        rev_per_case = ceil(float(nb_standard_reviewers * int(case_per_r)) / float(nb_cases))
    
    if method == 'percent per reviewer':
        rev_per_case = ceil( float( nb_standard_reviewers * int( ceil( float( nb_cases * int( percentage ) ) / float(100) ) ) ) / float(nb_cases) )
    
    return rev_per_case
 
def compute_case_per_rev(full, method, case_per_r, percentage):
    nb_cases = count_all_cases()
    nb_standard_reviewers = count_all_reviewers(False)
    
    if full or method == 'all for all' or nb_standard_reviewers == 0:
        case_per_rev = nb_cases
 
    if method == 'n per reviewer':
        case_per_rev = case_per_r
    
    if method == 'percent per reviewer':
        case_per_rev = ceil( float( nb_cases * int( percentage ) ) / float(100) )
    
    return case_per_rev

def get_data_for_export():

    wb = Workbook()
    ws = wb.active
    ws.title = "Study_data"
    
    with open(os.path.join(getcwd(), 'persistence', 'study_name.txt'), 'r', encoding="utf-8") as fr:
        study_name = (fr.read()).replace('\n', '')
    
    file_name = study_name + '_study_data.xlsx'

    wb_path = os.path.join(getcwd(), 'results', file_name)
    
    finalExtract = extract_all_data()
    
    nbCategories = len(finalExtract[0].categories)
    
    gold_standard_existing = gold_standard_exists()
    
    one_of_category = 2
    column_increment = 1
    
    ws.cell(row=1, column=1, value='cases')
    ws.cell(row=1, column=2, value='reviewers')
    
    column_increment += 2
    
    for i in range(nbCategories):
        current_category = finalExtract[0].categories[i]
        if current_category.catType == one_of_category:
            ws.cell(row=1, column=column_increment, value=format_r_friendly(current_category.name))
            column_increment += 1
        else:
            for criterion in current_category.criteria:
                ws.cell(row=1, column=column_increment, value=format_r_friendly(criterion.name))
                column_increment += 1
        if current_category.confidence != -2:
            ws.cell(row=1, column=column_increment, value=format_r_friendly(format_r_friendly(current_category.name) + '_confidence'))
            column_increment += 1
            
    if gold_standard_existing:
        ws.cell(row=1 , column=column_increment, value='reviewer_gold_standard_answer')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='gold_standard_confidence')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='gold_standard_answer')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='gold_standard_comparison')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='reviewer_gold_standard_malignancy')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='gold_standard_malignancy')
        column_increment +=1
        ws.cell(row=1 , column=column_increment, value='gold_standard_malignancy_comparison')
        column_increment +=1
    
    for i in range(len(finalExtract)):
        column_increment = 1
        ws.cell(row=i + 2, column=1, value=finalExtract[i].case)
        ws.cell(row=i + 2, column=2, value=finalExtract[i].reviewer)
        column_increment +=2
        for current_category in finalExtract[i].categories:
            if current_category.catType == one_of_category:
                ws.cell(row=i + 2, column=column_increment, value=format_r_friendly(current_category.diagnosis))
                column_increment += 1
                
            else:
                for criterion in current_category.criteria:
                    ws.cell(row=i + 2, column=column_increment, value=criterion.value)
                    column_increment += 1
                    
            if current_category.confidence != -2:
                ws.cell(row=i + 2, column=column_increment, value=current_category.confidence)
                column_increment += 1
        
        if gold_standard_existing:
            ws.cell(row=i + 2, column=column_increment, value=format_r_friendly(finalExtract[i].reviewer_gold_standard_answer))
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=finalExtract[i].gold_standard_confidence)
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=format_r_friendly(finalExtract[i].gold_standard_answer))
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=finalExtract[i].gold_standard_comparison)
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=format_r_friendly(finalExtract[i].reviewer_gold_standard_malignancy))
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=format_r_friendly(finalExtract[i].gold_standard_malignancy))
            column_increment +=1
            ws.cell(row=i + 2, column=column_increment, value=finalExtract[i].gold_standard_malignancy_comparison)
            column_increment +=1
        
    wb.save(wb_path)
    
    return (wb_path, file_name)
    
def clear_data():

    clear_all_answers()
    clear_all_cases()
    clear_all_prerequisites()
    clear_all_criteria()
    clear_all_categories()
    clear_non_admin_users()

    remove_case_images()

    if os.path.exists(os.path.join(getcwd(), 'uploads', 'tutorial_images.zip')):
        remove_tutorial_images()
    
    if os.path.exists(os.path.join(getcwd(), 'uploads', 'case_data')):
        remove_case_data()

def delete_user(userId):
    delete_reviewer_by_id(userId)
    
def regenerate_password(userId):
    newPass = generate_password()
    s = gensalt()
    hashPass = hashpw(newPass.encode('utf-8'), s).decode('utf-8')
    update_password(userId, hashPass)
    
    return newPass

