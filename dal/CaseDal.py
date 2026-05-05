#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Aug 27 20:23:30 2025

@author: jacques
"""

#general imports
from sqlalchemy import select
from sqlalchemy import delete

from sqlalchemy.orm import Session

from os import getcwd
from os import listdir
import os.path

from openpyxl import load_workbook

#enable imports from local modules
from pathlib import Path
import sys
path_root = Path(__file__).parents[2]
sys.path.append(str(path_root))

#local imports
from img_vote.utilities.useful import natural_sort_key, format_r_friendly

from img_vote.Models.DataModels import CaseDataModel, FinalExtractDataModel, CategoryExtractDataModel, CriterionExtractdataModel
from img_vote.Models.POCO import CasePOCO, CriterionPOCO, AnswerPOCO, AnswerCriterionPOCO, CategoryPOCO
from img_vote.dal.CriterionDal import get_gold_standard_criteria
from img_vote.dal.UserDal import get_all_non_admin_reviewers


#read only     
def get_case_by_id(identifier, engine):
    
    session = Session(engine)
    
    try:
        casePOCO = session.get(CasePOCO, identifier)
        
        case = CaseDataModel(casePOCO.id, casePOCO.path, casePOCO.name)

    finally:        
        session.close()
    
    return case

def get_all_cases(engine):
    
    session = Session(engine)
    
    cases = []
    
    try:
        casesQuery = select(CasePOCO)
        casesPOCO = session.execute(casesQuery).all()
        
        for i in range(len(casesPOCO)):
            cases.append(CaseDataModel(casesPOCO[i][0].id, casesPOCO[i][0].path, casesPOCO[i][0].name)) #for some reason all() returns a tuple, interesting value in first place, see later
    
    finally:
        session.close()
    
    return cases

def count_all_cases(engine):
    
    session = Session(engine)

    try:
        query = session.query(CasePOCO)
        answer = query.count()
    
    finally:
        session.close()
    
    return answer
    
def exists_case_by_id(identifier, engine):
    
    session = Session(engine)
    
    try:
        query = session.query(CasePOCO).filter(CasePOCO.id == identifier)
        ans = session.query(query.exists()).scalar()
    
    finally:
        session.close()
    
    return ans

#CRUD
def create_case(path, name, gld_std, engine):
    
    session = Session(engine)
    
    try:
        newCase= CasePOCO(path, name, gld_std)
        
        session.add(newCase)
        session.commit()
        
        caseId = newCase.id

    finally:        
        session.close()
    
    return caseId
    
#final data extraction
def extract_all_data(engine):
    
    session = Session(engine)
    
    one_of_category = 2
    trueValue = 1
    
    reviewers = get_all_non_admin_reviewers(engine)
    
    try:
        gold_standard_category = session.query(CategoryPOCO).filter(CategoryPOCO.has_gold_standard == True).one_or_none()
        if gold_standard_category != None:
            casesQuery = session.query(CasePOCO, CriterionPOCO).join(CriterionPOCO, CasePOCO.gold_standard == CriterionPOCO.id).order_by(CasePOCO.id)
        else:
            casesQuery = session.query(CasePOCO).order_by(CasePOCO.id)
            
        cases = session.execute(casesQuery).all()
        categories = session.query(CategoryPOCO).filter(CategoryPOCO.has_gold_standard == False).order_by(CategoryPOCO.id).all()
        
        extracts = []
        
        for case in cases:
            for reviewer in reviewers:
                
                answerExistsQuery = session.query(AnswerPOCO).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId)
                answerExists = session.query(answerExistsQuery.exists()).scalar()
                
                #check first if reviewer was assigned to this case
                if answerExists:
                    
                    rev_identifier = str(reviewer.userId) + '_' + format_r_friendly(reviewer.name)
                    
                    newExtract = FinalExtractDataModel(case[0].id, rev_identifier)
                    
                    newExtract.categories = []
                    
                    if gold_standard_category != None:
                        gold_standard_query = session.query(AnswerCriterionPOCO, CriterionPOCO, AnswerPOCO).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId).where(CriterionPOCO.category == gold_standard_category.id).where(AnswerCriterionPOCO.value == trueValue).where(CriterionPOCO.is_trust == False)
                        gold_standard_answer = gold_standard_query.one_or_none()
                        
                        
                        if gold_standard_answer != None:
    
                            rev_diag = 'unanswered'
                                
                            rev_diag = format_r_friendly(gold_standard_answer[1].name)
                            
                            gld_std_name = format_r_friendly(case[1].name)
                            
                            newExtract.reviewer_gold_standard_answer = rev_diag
                            newExtract.gold_standard_answer = gld_std_name
                            newExtract.gold_standard_comparison = rev_diag == gld_std_name
    
                            if gold_standard_category.has_malignancy:
    
                                rev_malignancy = 'unanswered'
                                gld_std_malignancy = 'unanswered'
    
                                if gold_standard_answer[1].malignancy:
                                    rev_malignancy = 'malignant'
                                else:
                                    rev_malignancy = 'benign'
        
                                if case[1].malignancy:
                                    gld_std_malignancy = 'malignant'
                                else:            
                                    gld_std_malignancy = 'benign'
    
                                newExtract.reviewer_gold_standard_malignancy = rev_malignancy
                                newExtract.gold_standard_malignancy = gld_std_malignancy
                                newExtract.gold_standard_malignancy_comparison = rev_malignancy == gld_std_malignancy
                            
                            if gold_standard_category.has_trust:
                            
                                gold_standard_trust_query = session.query(AnswerCriterionPOCO, CriterionPOCO, AnswerPOCO).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId).where(CriterionPOCO.category == gold_standard_category.id).where(CriterionPOCO.is_trust == True)
                                gold_standard_trust_answer = gold_standard_trust_query.one_or_none()
    
                                newExtract.gold_standard_confidence = gold_standard_trust_answer[0].value
                            
                        
                    for category in categories:
                        
                        newCategoryDM = CategoryExtractDataModel(category.name, category.type)
                        
                        newCategoryDM.criteria = []
                        
                        if category.type != one_of_category:
                            queryCriteria = session.query(AnswerCriterionPOCO, CriterionPOCO, AnswerPOCO).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId).where(CriterionPOCO.category == category.id).where(CriterionPOCO.is_trust == False).order_by(CriterionPOCO.id)
                            ansCriteria = queryCriteria.all()
                            
                            for ansCriterion in ansCriteria:
                                newCategoryDM.criteria.append(CriterionExtractdataModel(ansCriterion[1].name, ansCriterion[0].value))
                        
                        else:
                            queryDiagnosis = session.query(AnswerCriterionPOCO, CriterionPOCO, AnswerPOCO).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId).where(CriterionPOCO.category == category.id).where(AnswerCriterionPOCO.value == trueValue).where(CriterionPOCO.is_trust == False)
                            ansDiagnosis = queryDiagnosis.one_or_none()
                            if ansDiagnosis != None:
                                newCategoryDM.diagnosis = ansDiagnosis[1].name
                            else:
                                newCategoryDM.diagnosis = 'unanswered'
                                
                        
                        if category.has_trust:
                            queryConfidence = session.query(AnswerCriterionPOCO, CriterionPOCO, AnswerPOCO).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).where(AnswerPOCO.study_case == case[0].id).where(AnswerPOCO.reviewer == reviewer.userId).where(CriterionPOCO.category == category.id).where(CriterionPOCO.is_trust == True)
                            ansConfidence = queryConfidence.one_or_none()
                            newCategoryDM.confidence = ansConfidence[0].value
                        
                        newExtract.categories.append(newCategoryDM)
                    
                    extracts.append(newExtract)
    
    finally:
        session.close()
        
    return extracts

#One-time data creation
def create_all_cases(engine):
    
    session = Session(engine)

    try:
        criteria = get_gold_standard_criteria(engine)
        gold_standard = len(criteria) > 0
        
        if gold_standard:
            criteriaDict = dict()
            
            for criterion in criteria:
                criteriaDict[criterion[1]] = criterion[0]
        
        cwd = getcwd()
        path = os.path.join(cwd,'data', 'Img_data')
        case_files = [file for file in listdir(path) if not file.startswith('.')]
        
        if gold_standard:
            wbfolderpath = os.path.join(cwd, 'data')
            for filename in os.listdir(wbfolderpath):
                if filename.startswith('case_data'):
                    wbpath = os.path.join(wbfolderpath, filename)
        
            case_files = sorted(list(case_files), key=natural_sort_key)
            wb_obj = load_workbook(wbpath)
            sheet_obj = wb_obj.active
        
        counter = 2
        
        for case_file in case_files:
            newpath = os.path.join(path, case_file)
            
            if gold_standard:
                newname = str(sheet_obj.cell(row=counter, column=1).value)

                if newname != case_file:
                    session.rollback()
                    session.close()
                    
                    return ('file name discrepancy', filename, newname)

                cell_value = sheet_obj.cell(row=counter, column=2).value
                gld_std_name = cell_value.removesuffix(' ')

                try:
                    gld_std = criteriaDict[gld_std_name]
                except KeyError:
                    return ('gold standard name discrepancy', gld_std_name)
                newcase = CasePOCO(newpath, newname, gld_std)
            else:
                newname = str(counter - 1)
                newcase = CasePOCO(newpath, newname)
                
            counter += 1
            
            session.add(newcase)
        
        session.commit()
        
    finally: 
        session.close()
     
    return None
   
def clear_all_cases(engine):
    
    session = Session(engine)
    
    try:    
        deleteStmt = delete(CasePOCO)
        
        session.execute(deleteStmt)
        session.commit()
    
    finally:
        session.close()
