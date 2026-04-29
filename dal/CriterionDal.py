#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Aug 28 16:15:24 2025

@author: jacques
"""

#general imports
from os import getcwd
import os.path


from openpyxl import load_workbook

from sqlalchemy import select
from sqlalchemy import update
from sqlalchemy import delete

from sqlalchemy.orm import Session

#enable imports from local modules
from pathlib import Path
import sys
path_root = Path(__file__).parents[2]
sys.path.append(str(path_root))

#local imports
from img_vote.Models.DataModels import CriterionDataModel, CriterionForCaseDataModel, CategoryWithCriteriaDataModel, CategoryWithCriteriaAndPrerequisitesDataModel, DiagnosisDataModel, FinalExtractDataModel
from img_vote.Models.POCO import ReviewerPOCO, CasePOCO, AnswerPOCO, CriterionPOCO, AnswerCriterionPOCO, CategoryPOCO, PrerequisitePOCO
from img_vote.dal.AnswerDal import get_all_answers
from img_vote.dal.AnswerDal import get_user_answers

     
#read-only 
def get_criterion_by_id(identifier, engine):
    
    session = Session(engine)
    
    criterionPOCO = session.get(CriterionPOCO, identifier)
    
    criterion = CriterionDataModel(criterionPOCO.id, criterionPOCO.name, criterionPOCO.tutorial_path, criterionPOCO.category, criterionPOCO.is_trust)
    
    session.close()
    
    return criterion

def is_answer_done(userId, caseId, engine):
    
    session = Session(engine)
    
    #criterion category 2 is one of
    yes_no_category = 1
    one_of_category = 2 
    unanswered_value = 0
    unanswered_trust_value = 0
    true_value = 1
    
    # prefiltered query with all necessary joins
    query0 = session.query(CasePOCO, AnswerPOCO, CriterionPOCO, AnswerCriterionPOCO, CategoryPOCO).join(AnswerPOCO, CasePOCO.id == AnswerPOCO.study_case).join(AnswerCriterionPOCO, AnswerPOCO.id == AnswerCriterionPOCO.answer).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).join(CategoryPOCO, CriterionPOCO.category == CategoryPOCO.id).filter(CasePOCO.id == caseId).filter(AnswerPOCO.reviewer == userId)
    
    categories = session.query(CategoryPOCO).all()
    #0: case, 1:answer, 2: criterion, 3:answercriterion, 4: category
    for category in categories:
        
        mandatory = True
        
        if category.optional:
            
            prerequisites = session.query(PrerequisitePOCO.criterion).filter(PrerequisitePOCO.category == category.id).all()
            #check if all prerequisites are checked
            for prerequisite in prerequisites:
                prerequisitesQuery = query0.filter(CriterionPOCO.id == prerequisite.criterion).filter(AnswerCriterionPOCO.value != true_value)
                if session.query(prerequisitesQuery.exists()).scalar():
                    mandatory = False
                    
        if mandatory:
            if category.type == yes_no_category:
                #check if any criterion is still unanswered
                unanswered_query = query0.filter(CategoryPOCO.id == category.id).filter(AnswerCriterionPOCO.value == unanswered_value).filter(CriterionPOCO.is_trust == False)
                if session.query(unanswered_query.exists()).scalar():
                    return False
            
            if category.type == one_of_category:
                #check if any criterion has been chosen
                answered_query = query0.filter(CategoryPOCO.id == category.id).filter(AnswerCriterionPOCO.value == true_value).filter(CriterionPOCO.is_trust == False)
                if not session.query(answered_query.exists()).scalar():
                    return False
            
            if category.has_trust:
                #check if trust criterion was left unanswered
                trust_query = query0.filter(CategoryPOCO.id == category.id).filter(AnswerCriterionPOCO.value == unanswered_trust_value).filter(CriterionPOCO.is_trust == True)
                if session.query(trust_query.exists()).scalar():
                    return False
        
    return True

def get_criteria_for_case(userId, caseId, catId, engine):

    session = Session(engine)

    try:
        query = session.query(CasePOCO, AnswerPOCO, CriterionPOCO, AnswerCriterionPOCO).join(AnswerPOCO, CasePOCO.id == AnswerPOCO.study_case).join(AnswerCriterionPOCO, AnswerPOCO.id == AnswerCriterionPOCO.answer).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).filter(CriterionPOCO.category == catId).filter(AnswerPOCO.reviewer == userId).filter(CasePOCO.id == caseId).order_by(CriterionPOCO.category, CriterionPOCO.id)
    
        queriedAnswers = query.all()
    
    
        answer = []
    
        # 0: Case, 1: Answer, 2: Criterion, 3: AnswerCriterion
        for queriedAnswer in queriedAnswers:
            answer.append(CriterionForCaseDataModel(queriedAnswer[2].id, queriedAnswer[2].name, queriedAnswer[3].value, queriedAnswer[2].category, queriedAnswer[2].is_trust, queriedAnswer[2].tutorial_path))
    
    finally:
        session.close()

    return answer


def get_diagnosis_for_case(userId, caseId, engine):
    
    session = Session(engine)
    #criterion type 2 is diagnosis
    diagnosis_type = 2    

    query = session.query(CasePOCO, AnswerPOCO, CriterionPOCO, AnswerCriterionPOCO).join(AnswerPOCO, CasePOCO.id == AnswerPOCO.study_case).join(AnswerCriterionPOCO, AnswerPOCO.id == AnswerCriterionPOCO.answer).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).filter(CasePOCO.id == caseId).filter(AnswerPOCO.reviewer == userId).filter(CriterionPOCO.type == diagnosis_type).order_by(CriterionPOCO.id)

    queriedAnswer = query.all()
    
    #0: Case, 1: Answer, 2: Criterion, 3: AnswerCriterion   
    
    answer = CriterionForCaseDataModel(queriedAnswer[0][0].path)
    
    for i in range(len(queriedAnswer)):
        answer.criteria.append((queriedAnswer[i][2].type, queriedAnswer[i][2].name, queriedAnswer[i][3].value, queriedAnswer[i][2].tutorial_path, queriedAnswer[i][2].id))
    
    session.close()
    
    return answer

def get_all_criteria(engine):
    
    session = Session(engine)

    try:
        criteria = []
        
        criteriaQuery = select(CriterionPOCO)
        criteriaPOCO = session.execute(criteriaQuery).all()
        
        for i in range(len(criteriaPOCO)):
            criteria.append(criteriaPOCO[i][0]) #no time to investigate all(), mayhap later

    finally:
        session.close()

    return criteria

def get_all_criteria_no_diagnosis(engine):
    
    session = Session(engine)
    
    diagnosis_type = 0
    trust_category = 3
    
    query = session.query(CriterionPOCO).filter(CriterionPOCO.type != diagnosis_type).filter(CriterionPOCO.category != trust_category)
    ans = query.all()
    
    critNoDiag = []
    
    for i in range(len(ans)):
        critNoDiag.append(DiagnosisDataModel(ans[i].id, ans[i].name))
    
    session.close()
    
    return critNoDiag

def get_gold_standard_criteria(engine):
    
    session = Session(engine)

    try:
        query = session.query(CriterionPOCO.id, CriterionPOCO.name).join(CategoryPOCO, CriterionPOCO.category == CategoryPOCO.id).filter(CategoryPOCO.has_gold_standard == True)
        answer = query.all()
    finally:
        session.close()
    
    return answer

def categories_with_criteria(engine):
    
    session = Session(engine)
    
    query = session.query(CategoryPOCO, CriterionPOCO).outerjoin(CriterionPOCO, CategoryPOCO.id == CriterionPOCO.category).order_by(CategoryPOCO.id, CriterionPOCO.id)
    
    #0: category, 1:criterion
    queriedAnswer = query.all()
    
    answer = []
    currentCategory = -1
    currentDataModel = None
    
    for ans in queriedAnswer:
        if (ans[0].id != currentCategory):
            if (currentCategory != -1):
                answer.append(currentDataModel)
            currentCategory = ans[0].id
            currentDataModel = CategoryWithCriteriaDataModel(ans[0].id, ans[0].name, ans[0].type, ans[0].has_tutorial, ans[0].has_trust, ans[0].has_na, ans[0].optional)
        if ans[1] != None and not ans[1].is_trust:
            currentDataModel.criteria.append((ans[1].id, ans[1].name))
    
    #avoid off by one
    if currentDataModel != None:
        answer.append(currentDataModel)
    
    return answer

def category_with_criteria_and_prerequisites(catId, engine):
    
    session = Session(engine)
    
    query1 = session.query(CategoryPOCO, CriterionPOCO).outerjoin(CriterionPOCO, CategoryPOCO.id == CriterionPOCO.category).filter(CategoryPOCO.id == catId).order_by(CategoryPOCO.id, CriterionPOCO.id)
    query2 = session.query(CriterionPOCO, PrerequisitePOCO).join(PrerequisitePOCO, CriterionPOCO.id == PrerequisitePOCO.criterion).filter(PrerequisitePOCO.category == catId)
    
    #0: category, 1: criterion
    queriedAnswer1 = query1.all()
    #0: criterion, 1: prerequisite
    queriedAnswer2 = query2.all()
    
    
    answer = queriedAnswer1[0]
    categoryDataModel = CategoryWithCriteriaAndPrerequisitesDataModel(answer[0].id, answer[0].name, answer[0].type, answer[0].has_tutorial, answer[0].has_trust, answer[0].has_na, answer[0].optional, answer[0].has_gold_standard, answer[0].has_malignancy)
    
    for ans in queriedAnswer1:
        if ans[1] != None and not ans[1].is_trust:
            categoryDataModel.criteria.append((ans[1].id, ans[1].name, ans[1].malignancy))
    
    for ans in queriedAnswer2:
        categoryDataModel.prerequisites.append((ans[0].id, ans[0].name))
    
    
    return categoryDataModel

#CRUD
def create_criterion(name, tutorial_path, category, is_trust, malignancy, engine):
    
    session = Session(engine)

    newCrit = CriterionPOCO(name, tutorial_path, category, is_trust, malignancy)
    if (newCrit.tutorial_path == None):
        newCrit.tutorial_path = tutorial_path
        
    session.add(newCrit)
    session.commit()
    
    session.close()

def update_criterion(crit_id, crit_name, crit_malignancy, engine):
    
    session = Session(engine)
    
    updatestmt = update(CriterionPOCO).where(CriterionPOCO.id == crit_id).values(name=crit_name, malignancy=crit_malignancy)

    session.execute(updatestmt)

    session.commit()
    
    session.close()
    
def update_criterion_malignancy(crit_id, crit_malignancy, engine):
    
    session = Session(engine)
    
    updatestmt = update(CriterionPOCO).where(CriterionPOCO.id == crit_id).values(malignancy=crit_malignancy)

    session.execute(updatestmt)

    session.commit()
    
    session.close()

def update_criteria_path(path, engine):
    
    session = Session(engine)
    
    try:    
        no_tutorial = session.query(CriterionPOCO.id).join(CategoryPOCO, CriterionPOCO.category == CategoryPOCO.id).filter(CategoryPOCO.has_tutorial == False).all()
        
        for no_tuto in no_tutorial:
        
            updatestmt = update(CriterionPOCO).where(CriterionPOCO.id == no_tuto.id).values(tutorial_path='')
            
            session.execute(updatestmt)
        
        has_tutorial = session.query(CriterionPOCO, CategoryPOCO).join(CategoryPOCO, CriterionPOCO.category == CategoryPOCO.id).filter(CategoryPOCO.has_tutorial == True).all()
        
        #0: Criterion, 1: Category
        
        for answer in has_tutorial:
            path = os.path.join(path, answer[1].name, answer[0].name)
            updatestmt = update(CriterionPOCO).where(CriterionPOCO.id == answer[0].id).values(tutorial_path=path)
            session.execute(updatestmt)
    
    finally:
        session.close()

def clear_malignant_criteria_in_non_malignant_category(engine):
    
    session = Session(engine)
    
    try:
        
        query = session.query(CriterionPOCO, CategoryPOCO).join(CategoryPOCO, CriterionPOCO.category == CategoryPOCO.id).filter(CriterionPOCO.malignancy == True).filter(CategoryPOCO.has_malignancy == False)
        queriedAnswer = query.all()
        
        #0: criterion, 1: category
        
        for ans in queriedAnswer:
            updatestmt = update(CriterionPOCO).where(CriterionPOCO.id == ans[0].id).values(malignancy=False)
            session.execute(updatestmt)
        
        session.commit()
        
    finally:
        session.close()

def erase_criterion(identifier, engine):
    
    session = Session(engine)
    
    deleteStmt = delete(CriterionPOCO).where(CriterionPOCO.id == identifier)
    
    session.execute(deleteStmt)
    session.commit()

    session.close()
    
def erase_category_criteria(identifier, engine):
    
    session = Session(engine)
    
    deleteStmt = delete(CriterionPOCO).where(CriterionPOCO.category == identifier)
    
    session.execute(deleteStmt)
    session.commit()

    session.close()

def create_answer_to_criterion(answer, criterion, engine):
    
    session = Session(engine)
    
    newAnsCrit = AnswerCriterionPOCO(answer, criterion)
    session.add(newAnsCrit)
    session.commit()
    
    session.close()

def save_Criterion(user, case, critName, newValue, engine):
    
    session = Session(engine)
    
    query = session.query(AnswerCriterionPOCO, AnswerPOCO, CriterionPOCO).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).filter(CriterionPOCO.name == critName).filter(AnswerPOCO.study_case == case).filter(AnswerPOCO.reviewer == user)
    ansCrit = query.one_or_none()[0]
    updatestmt = update(AnswerCriterionPOCO).where(AnswerCriterionPOCO.answer == ansCrit.answer).where(AnswerCriterionPOCO.criterion == ansCrit.criterion).values(value=newValue)

    session.execute(updatestmt)

    session.commit()
    
    session.close()

def safeguard_Criterion(user, case, critId, newValue, engine):
    
    session = Session(engine)
     
    try:
        query = session.query(AnswerCriterionPOCO, AnswerPOCO, CriterionPOCO).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).filter(CriterionPOCO.id == critId).filter(AnswerPOCO.study_case == case).filter(AnswerPOCO.reviewer == user)
        
        result = query.one_or_none()
        
        if result is None:
            return
        
        ansCrit = result[0]
                
        updatestmt = update(AnswerCriterionPOCO).where(AnswerCriterionPOCO.answer == ansCrit.answer).where(AnswerCriterionPOCO.criterion == ansCrit.criterion).values(value=newValue)
    
        session.execute(updatestmt)
    
        session.commit()
        
    finally:
        session.close()


def undo_all_but_one(userId, case, criterionId, value, category, engine):
    
    session = Session(engine)
    
    try:
        true_value = 1
        false_value = 2
        
        query = session.query(AnswerCriterionPOCO, AnswerPOCO, CriterionPOCO).join(AnswerPOCO, AnswerCriterionPOCO.answer == AnswerPOCO.id).join(CriterionPOCO, AnswerCriterionPOCO.criterion == CriterionPOCO.id).filter(CriterionPOCO.category == category).filter(AnswerPOCO.reviewer == userId).filter(AnswerPOCO.study_case == case).filter(CriterionPOCO.id != criterionId).filter(CriterionPOCO.is_trust == False)
        ansCrit = query.all()
    
        # 0: answerCriterion, 1: answer, 2: criterion
        ansId = ansCrit[0][1].id
        
        for ans in ansCrit:
            
            updatestmt = update(AnswerCriterionPOCO).where(AnswerCriterionPOCO.answer == ansId).where(AnswerCriterionPOCO.criterion == ans[2].id).values(value=false_value)
            session.execute(updatestmt)
        
        
        updatestmt = update(AnswerCriterionPOCO).where(AnswerCriterionPOCO.answer == ansId).where(AnswerCriterionPOCO.criterion == criterionId).values(value=value)
        session.execute(updatestmt)
        
        session.commit()
     
    finally:
        session.close()
    
    
#one-time data creation
def create_trust_criteria(engine):
    
    session = Session(engine)
    
    try:
        query = session.query(CategoryPOCO).filter(CategoryPOCO.has_trust == True).order_by(CategoryPOCO.id)
        
        answer = query.all()
        
        for ans in answer:
            newCrit = CriterionPOCO(ans.name + '_trust_scale', '', ans.id, True, False)
                
            if (newCrit.tutorial_path == None):
                newCrit.tutorial_path = ''
                
            session.add(newCrit)
            
        session.commit()
        
    finally:
        session.close()

def create_na_criteria(engine):
    
    session = Session(engine)
    
    one_of_category = 2
    
    try:
        query = session.query(CategoryPOCO).filter(CategoryPOCO.has_na == True).filter(CategoryPOCO.type == one_of_category).order_by(CategoryPOCO.id)
        
        answer = query.all()
        
        for ans in answer:
            newCrit = CriterionPOCO('na', '', ans.id, False, False)
                
            if (newCrit.tutorial_path == None):
                newCrit.tutorial_path = ''
                
            session.add(newCrit)
            
        session.commit()
        
    finally:
        session.close()

def create_all_criterion(engine):
    
    # load delimiting words from config file
    delimitations = []
    with open(os.path.join(getcwd(), 'persistence', 'delimitations.txt'), encoding="utf-8") as delim_file:
        for l in delim_file:
            delimitations.append(l.removesuffix('\n'))
          
    critType = -1
    trust = False
    name = ""
    typeName = ""
    
    criteriaWB = load_workbook(filename = os.path.join(getcwd(), 'data', 'WOW.xlsx'))
    criteriaWS = criteriaWB.active
    
    session = Session(engine)
    
    try:
        for i in range (1, criteriaWS.max_row + 1):
            nameCell = criteriaWS.cell(row=i, column=1)
            categoryCell = criteriaWS.cell(row=i, column=2)
            trustCell = criteriaWS.cell(row=i, column=3)
            malCell = criteriaWS.cell(row=i, column=4)
            if nameCell.value in delimitations:
                
                # add trust criterion at the end of a type
                if critType > -1 and trust:
                    newCrit = CriterionPOCO(typeName, "", critType, 3, False)
                    newCrit.tutorial_path = ""
                    session.add(newCrit)
                
                typeName = nameCell.value
    
                critType +=1
                categoryName = categoryCell.value
                trust = trustCell.value == "With Trust"
                if categoryName == "One Of":
                    category = 2
                if categoryName == "Multiple":
                    category = 1
               
            else:
                #dirty fix, remove later
                name = nameCell.value
                tutorial_name = name
                if tutorial_name.find(">") != -1:
                    tutorial_name = name.replace(">", "gt ")
                tutorial_slide_path = os.path.join(getcwd(), 'data', 'tutorial_data', delimitations[critType], tutorial_name + '.png')
                
                malignancy = (malCell.value == "Malignant")
                newCrit = CriterionPOCO(name, tutorial_slide_path, critType, category, malignancy)
                newCrit.tutorial_path = tutorial_slide_path
                session.add(newCrit)
            
            # add trust criterion at the end if need be
            if (i == criteriaWS.max_row) and trust:
                newCrit = CriterionPOCO(typeName, "", critType, 3, False)
                newCrit.tutorial_path = ""
                session.add(newCrit) 
        
        session.commit()
    
    finally:
        session.close()

def create_all_answer_to_criterion(engine):
        
    answers = get_all_answers(engine)
    criteria = get_all_criteria(engine)
    
    session = Session(engine)
    
    try:
        for i in range(len(answers)):
            for j in range(len(criteria)):
                newAnsCrit = AnswerCriterionPOCO(answers[i].id, criteria[j].id)
                session.add(newAnsCrit)
    
        session.commit()

    finally:
        session.close()

def create_user_answer_to_criterion(userId, engine):
        
    answers = get_user_answers(userId, engine)
    criteria = get_all_criteria(engine)
    
    session = Session(engine)
    
    for i in range(len(answers)):
        for j in range(len(criteria)):
            newAnsCrit = AnswerCriterionPOCO(answers[i].id, criteria[j].id)
            session.add(newAnsCrit)
    
    session.commit()
    
    session.close()

def clear_all_criteria(engine):
    
    session = Session(engine)
    
    try:
        deleteStmt = delete(CriterionPOCO)
        
        session.execute(deleteStmt)
        session.commit()
    
    finally:
        session.close()
    
